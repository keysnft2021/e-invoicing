"""Bulk upload — accepts CSV or XLSX, parses rows and creates invoices.

CSV columns (case-insensitive):
    customer_tin, customer_name, invoice_date (YYYY-MM-DD), currency,
    description, quantity, unit_price, tax_rate, business_system, store_code

One row = one line item on an invoice. Rows with the same customer_tin +
invoice_date within an upload are grouped into a single invoice.
"""
import csv
import io
from datetime import datetime, timezone
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

from deps import get_db, require_tenant
from audit import audit

router = APIRouter(prefix="/api/ics/bulk", tags=["bulk"])


def _next_number(prefix: str) -> str:
    dt = datetime.now(timezone.utc)
    return f"{prefix}-{dt.strftime('%Y%m')}-{int(dt.timestamp() * 1000) % 100000:05d}"


def _parse_csv(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [{(k or "").strip().lower(): (v.strip() if v else "") for k, v in row.items()}
             for row in reader]


def _parse_xlsx(raw: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(400, "openpyxl not installed")
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        out.append({headers[i]: (str(v).strip() if v is not None else "")
                    for i, v in enumerate(r) if i < len(headers)})
    return out


@router.get("/template")
async def download_template():
    """Sample CSV template."""
    sample = (
        "customer_tin,customer_name,invoice_date,currency,description,quantity,unit_price,tax_rate,business_system,store_code\n"
        "C11112223334,Global Retail Sdn Bhd,2026-02-15,MYR,Consulting hours,10,350.00,6,SAP-ECC,HQ-KL01\n"
        "C11112223334,Global Retail Sdn Bhd,2026-02-15,MYR,License,1,4800.00,8,SAP-ECC,HQ-KL01\n"
        "IG55667788990,Sinar Cahaya Enterprise,2026-02-16,MYR,Steel sheet,5,1200.00,6,SAP-ECC,PG-01\n"
    )
    return StreamingResponse(
        io.BytesIO(sample.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ics-bulk-invoice-template.csv"},
    )


@router.post("/upload")
async def upload(file: UploadFile = File(...), ctx=Depends(require_tenant)):
    raw = await file.read()
    name = (file.filename or "").lower()
    try:
        rows = _parse_xlsx(raw) if name.endswith(".xlsx") else _parse_csv(raw)
    except Exception as e:
        raise HTTPException(400, f"Parse error: {type(e).__name__}: {e}")

    db = get_db()
    now = datetime.now(timezone.utc)
    tenant_id = ctx["tenant_id"]

    # Group rows by (customer_tin, invoice_date)
    groups: dict[tuple, list[dict]] = {}
    parse_errors: list[dict] = []
    for i, r in enumerate(rows, start=2):  # header is row 1
        try:
            key = (r.get("customer_tin", ""), r.get("invoice_date", ""))
            if not key[0] or not key[1]:
                parse_errors.append({"row": i, "error": "customer_tin and invoice_date are required"})
                continue
            groups.setdefault(key, []).append({
                "description": r.get("description", "") or "Item",
                "quantity": float(r.get("quantity") or 1),
                "unit_price": float(r.get("unit_price") or 0),
                "tax_rate": float(r.get("tax_rate") or 6),
                "discount": 0,
                "hs_code": r.get("hs_code") or None,
            })
        except Exception as e:
            parse_errors.append({"row": i, "error": f"{type(e).__name__}: {e}"})

    invoices_created: list[str] = []
    for (cust_tin, inv_date), lines in groups.items():
        # Find or create customer by TIN
        first_row = next((r for r in rows if r.get("customer_tin") == cust_tin), {})
        cust = await db.customers.find_one({"tenant_id": tenant_id, "tin": cust_tin})
        if not cust:
            cust_res = await db.customers.insert_one({
                "tenant_id": tenant_id,
                "name": first_row.get("customer_name") or cust_tin,
                "tin": cust_tin,
                "country": "MY", "currency": first_row.get("currency") or "MYR",
                "credit_limit": 0,
                "created_at": now.isoformat(),
            })
            cust = {"_id": cust_res.inserted_id, "name": first_row.get("customer_name") or cust_tin,
                     "tin": cust_tin}
        subtotal = sum((l["quantity"] * l["unit_price"] - l["discount"]) for l in lines)
        tax = sum((l["quantity"] * l["unit_price"] - l["discount"]) * (l["tax_rate"] / 100) for l in lines)
        total = round(subtotal + tax, 2)
        doc = {
            "tenant_id": tenant_id,
            "invoice_number": _next_number("INV"),
            "invoice_type": "invoice",
            "invoice_date": inv_date,
            "currency": first_row.get("currency") or "MYR",
            "customer_id": str(cust["_id"]),
            "customer_snapshot": {"id": str(cust["_id"]),
                                    "name": cust.get("name"),
                                    "tin": cust.get("tin")},
            "supplier_tin": None, "supplier_name": None,
            "business_system": first_row.get("business_system") or None,
            "store_code": first_row.get("store_code") or None,
            "source": "csv_upload",
            "invoice_confirmation_status": "pending",
            "validation_result": "pending",
            "lines": lines,
            "shipping": 0, "charges": 0, "round_off": 0,
            "subtotal": round(subtotal, 2), "tax_total": round(tax, 2), "total": total,
            "status": "draft", "government": {},
            "timeline": [{"status": "draft", "note": f"Bulk-uploaded from {file.filename}",
                            "actor": ctx["user"]["email"], "at": now.isoformat()}],
            "created_at": now.isoformat(), "updated_at": now.isoformat(),
            "created_by": ctx["user"]["id"], "created_by_email": ctx["user"]["email"],
        }
        res = await db.invoices.insert_one(doc)
        invoices_created.append(str(res.inserted_id))

    job = {
        "tenant_id": tenant_id,
        "filename": file.filename,
        "size_bytes": len(raw),
        "row_count": len(rows),
        "invoices_created": len(invoices_created),
        "invoice_ids": invoices_created,
        "parse_errors": parse_errors,
        "status": "completed" if invoices_created else ("failed" if parse_errors else "empty"),
        "uploaded_by": ctx["user"]["email"],
        "uploaded_at": now.isoformat(),
    }
    res = await db.upload_jobs.insert_one(job)
    await audit(db, tenant_id=tenant_id, actor_id=ctx["user"]["id"],
                actor_email=ctx["user"]["email"], action="ics.bulk_upload",
                entity="upload_job", entity_id=str(res.inserted_id),
                meta={"filename": file.filename, "created": len(invoices_created)})
    job["id"] = str(res.inserted_id)
    job.pop("_id", None)
    return job


@router.get("/jobs")
async def list_jobs(ctx=Depends(require_tenant), limit: int = 100):
    """Uploaded Records — recent bulk upload jobs."""
    db = get_db()
    out = []
    async for j in db.upload_jobs.find({"tenant_id": ctx["tenant_id"]}).sort("uploaded_at", -1).limit(limit):
        j["id"] = str(j.pop("_id"))
        out.append(j)
    return out
